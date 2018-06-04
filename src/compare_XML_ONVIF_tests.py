"""
Created on 18 avr. 2018

@author: vauge
"""
import os
import xml.etree.ElementTree as etree
import xmlschema
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
MYGREEN = 'FF00C800'


class MalformedResultsFile(Exception):
    pass


class Test:
    """
    Class that represent a test with it's name, requirement level,
    result, time and steps details
    """

    def __init__(self, requirement_level, result, steps):
        self.requirement_level = requirement_level
        self.result = result
        self.steps = steps


class Step:
    """
    Class that represent a step with it's name and result
    """

    def __init__(self, name, result, message):
        self.name = name
        self.result = result
        self.message = message


def construct_steps(step_nodes):
    """
    Return an array of Step object given an array of "StepResult" nodes
    :param step_nodes: Array of Node object representing StepResult XML nodes
    :return: Array of Step object
    """
    steps = []

    # Run through all StepResult nodes
    for sn in step_nodes:
        name = sn.find('StepName').text
        result = sn.find('Status').text
        message = sn.find('Message').text

        steps.append(Step(name, result, message))

    return steps


def construct_tests(result_nodes):
    """
    Return a dictionary of Test object given an array of "TestResult" nodes
    :param result_nodes: Array of Node object representing TestResult XML nodes
    :return: Dictionary of Test object
    """
    results = {}

    # Run through all TestResult nodes
    for rn in result_nodes:
        name = rn.find('TestInfo').find('Name').text
        requirement_level = rn.find('TestInfo').find('RequirementLevel').text
        result = rn.find('Log').find('TestStatus').text
        step_nodes = rn.find('Log').find('Steps').findall('StepResult')

        results[name] = Test(requirement_level, result, construct_steps(step_nodes))

    return results


def construct_results(file_content):
    """
    Return a dictionary of Test object given an XML file path
    :param file_content: result file content
    :return: Dictionary of Test object
    """
    # Validate XML file with XML Schema
    my_schema = xmlschema.XMLSchema('ONVIF_Device_Test_Tool.xsd')
    if not my_schema.is_valid(file_content):
        raise MalformedResultsFile('')

    # Parse XML file content
    root_node = etree.fromstring(file_content)
    result_nodes = root_node.find('Results').findall('TestResult')

    return construct_tests(result_nodes)


def analyse_results(results_set):
    """
    Print some stats about a results dictionary
    :param results_set: Tuple containing the name of result file and the dictionary of tests
    :return: None
    """
    name = results_set[0]
    results = results_set[1]

    # Counters
    total_tests = len(results)
    total_mandatory_tests = 0
    passed_tests = 0
    failed_mandatory_tests = 0

    for test in results.values():
        # Update counters
        if test.result == 'Passed':
            passed_tests += 1

        if test.requirement_level == 'Must':
            total_mandatory_tests += 1
            if test.result == 'Failed':
                failed_mandatory_tests += 1

    print('==={}==='.format(name))
    print('Percentage of passed tests: ' + str(passed_tests / total_tests * 100) + '%')
    print('Percentage of failed mandatory tests: ' + str(failed_mandatory_tests / total_mandatory_tests * 100) + '%\n')


def compare_steps(wb, name, num, requirement_level, steps_set1, steps_set2):
    """
    Compare 2 steps array and print steps list to the workbook in a new worksheet
    :param wb: Output workbook
    :param name: Name of the parent test
    :param num: Test number
    :param requirement_level: Requirement level of the parent test
    :param steps_set1: Tuple containing the name of the first result file and the associated array of Step object
    :param steps_set2: Tuple containing the name of the second result file and the associated array of Step object
    :return: None
    """
    name1 = steps_set1[0]
    name2 = steps_set2[0]
    steps1 = steps_set1[1]
    steps2 = steps_set2[1]

    # Create new worksheet
    ws = wb.create_sheet(str(num))

    # maximum number of test between steps1 and steps2
    max_step_number = max([len(steps1), len(steps2)])

    # Headers
    ws['A1'] = '{} is {}'.format(name, 'mandatory' if requirement_level == 'Must' else 'optional')
    ws['A2'] = name1
    ws['B2'] = name2

    for i in range(max_step_number):
        if i < len(steps1):
            # Step name
            ws['A' + str(3 + i)] = steps1[i].name
            # Color in red if failed, otherwise in green
            ws['A' + str(3 + i)].font = Font(color=RED) if steps1[i].result == "Failed" else Font(color=MYGREEN)
            # Append error message if failed
            ws['A' + str(3 + i)].value += ': ' + steps1[i].message if steps1[i].result == "Failed" else ''

        if i < len(steps2):
            # Step name
            ws['B' + str(3 + i)] = steps2[i].name
            # Color in red if failed, otherwise in green
            ws['B' + str(3 + i)].font = Font(color=RED) if steps2[i].result == "Failed" else Font(color=MYGREEN)
            # Append error message if failed
            ws['B' + str(3 + i)].value += ': ' + steps2[i].message if steps2[i].result == "Failed" else ''

    adjust_column_width(ws, 'A')
    adjust_column_width(ws, 'B')


def adjust_column_width(worksheet, column):
    """
    Update column width according to cells content length
    :param worksheet: Worksheet to update
    :param column: Column's name to update
    :return: None
    """
    worksheet.column_dimensions[column].width = max([len(c.value) if c.value else 0 for c in worksheet[column]])


def compare_results(results_set1, results_set2):
    """
    Compare 2 results dictionary and print result to a workbook
    :param results_set1: Tuple containing the name of the first result file and the associated dictionary of Test object
    :param results_set2: Tuple containing the name of the second result file and the associated dictionary of Test
    object
    :return: None
    """
    name1 = results_set1[0]
    name2 = results_set2[0]
    results1 = results_set1[1]
    results2 = results_set2[1]

    # Create a new workbook
    wb = Workbook()
    # Get the first worksheet
    ws = wb.active
    ws.title = 'Test differences'

    # Headers
    ws['B1'] = name1
    ws['C1'] = name2

    # Look for test result different in file 1 and 2
    i = 0
    for name in results1:
        if name not in results2:
            continue

        if results1[name].result != results2[name].result:
            ws['A' + str(2 + i)] = name
            ws['B' + str(2 + i)] = results1[name].result
            ws['C' + str(2 + i)] = results2[name].result

            # Update color
            ws['B' + str(2 + i)].font = Font(color=RED) if results1[name].result == "Failed" else Font(color=MYGREEN)
            ws['C' + str(2 + i)].font = Font(color=RED) if results2[name].result == "Failed" else Font(color=MYGREEN)

            # Compare step for this test and create a new associated worksheet
            compare_steps(wb, name, i + 2, results1[name].requirement_level, (name1, results1[name].steps),
                          (name2, results2[name].steps))
            i += 1

    # Test done in file 2 but not in file 1
    for name in results2:
        if name not in results1:
            ws['A' + str(2 + i)] = name
            ws['B' + str(2 + i)] = '-'
            ws['C' + str(2 + i)] = results2[name].result

            # Update color
            ws['B' + str(2 + i)].font = Font(color=RED)
            ws['C' + str(2 + i)].font = Font(color=RED) if results2[name].result == "Failed" else Font(color=MYGREEN)

            i += 1

    # Test done in file 1 but not in file 2
    for name in results1:
        if name not in results2:
            ws['A' + str(2 + i)] = name
            ws['B' + str(2 + i)] = results1[name].result
            ws['C' + str(2 + i)] = '-'

            # Update color
            ws['B' + str(2 + i)].font = Font(color=RED) if results1[name].result == "Failed" else Font(color=MYGREEN)
            ws['C' + str(2 + i)].font = Font(color=RED)

            i += 1

    adjust_column_width(ws, 'A')
    adjust_column_width(ws, 'B')
    adjust_column_width(ws, 'C')

    # Write the workbook on disk
    wb.save('{}_vs_{}.xlsx'.format(name1, name2))


def usage():
    print('Usage: {} <file 1> [<file 2>]'.format(sys.argv[0]))
    sys.exit(1)


def clean_invalid_xml_characters(file):
    """
    Clean invalid XML characters
    :param file: XML file path
    :return: String of file content cleaned
    """
    invalid_xml = re.compile(r'&#x[0-1]?[0-9a-eA-E];')
    new_content = ''

    with open(file, 'r') as f:
        for line in f:
            nline, count = invalid_xml.subn('', line)
            new_content += nline

    return new_content


def main():
    """ main """
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        usage()

    # If there are 2 files compare them
    if len(sys.argv) == 3:
        if not os.path.exists(sys.argv[1]):
            sys.stderr.write("{} doesn't exist".format(sys.argv[1]))
            usage()

        if not os.path.exists(sys.argv[2]):
            sys.stderr.write("{} doesn't exist".format(sys.argv[2]))
            usage()

        name1 = os.path.basename(sys.argv[1])
        name2 = os.path.basename(sys.argv[2])

        # Clean invalid XML characters before parsing
        print('Cleaning {}...'.format(name1))
        file_content1 = clean_invalid_xml_characters(sys.argv[1])
        print('Cleaning {}...'.format(name2))
        file_content2 = clean_invalid_xml_characters(sys.argv[2])

        print('Parsing {}...'.format(name1))
        results1 = construct_results(file_content1)
        print('Parsing {}...'.format(name2))
        results2 = construct_results(file_content2)

        analyse_results((name1, results1))
        analyse_results((name2, results2))

        compare_results((name1, results1), (name2, results2))
    else:
        # Else print some stats about the only file
        if not os.path.exists(sys.argv[1]):
            sys.stderr.write("{} doesn't exist".format(sys.argv[1]))
            usage()

        name = os.path.basename(sys.argv[1])

        # Clean invalid XML characters before parsing
        print('Cleaning {}...'.format(name))
        file_content = clean_invalid_xml_characters(sys.argv[1])

        print('Parsing {}...'.format(name))
        results = construct_results(file_content)
        analyse_results((name, results))


if __name__ == '__main__':
    main()
